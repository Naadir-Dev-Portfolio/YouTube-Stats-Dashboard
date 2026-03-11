import sys
import os
import re
import ssl
import logging
from googleapiclient.discovery import build
import pandas as pd
from datetime import datetime
import numpy as np

from PyQt6.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QLabel, QPushButton,
    QComboBox, QMessageBox, QPlainTextEdit, QHBoxLayout, QScrollArea,
    QSplitter, QLineEdit, QToolTip
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal, QDate
from PyQt6.QtGui import QPalette, QColor, QFont, QCursor, QPainter
from PyQt6.QtCharts import (
    QChart, QChartView, QLineSeries, QDateTimeAxis, QValueAxis, QLegend
)
from PyQt6.QtWidgets import QGraphicsDropShadowEffect

import openpyxl
from openpyxl.chart import LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font


# Configure logging
logging.basicConfig(
    level=logging.DEBUG,
    filename='youtube_dashboard.log',
    filemode='w',
    format='%(asctime)s - %(levelname)s - %(message)s'
)


# Function to read API key from api.txt
def get_api_key():
    """Reads the API key from the api.txt file."""
    try:
        with open('api.txt', 'r') as f:
            api_key = f.read().strip()
            if not api_key:
                raise ValueError("API key is empty.")
            return api_key
    except FileNotFoundError:
        logging.critical("api.txt file not found.")
        raise FileNotFoundError("api.txt file not found.")
    except Exception as e:
        logging.critical(f"Error reading api.txt: {e}")
        raise e


# Function to extract channel ID from URL
def extract_channel_id(youtube, url):
    """Extracts the channel ID from a YouTube channel URL."""
    try:
        if '/channel/' in url:
            # URL contains the channel ID directly
            channel_id = url.split('/channel/')[-1].split('/')[0]
            logging.debug(f"Extracted Channel ID from URL: {channel_id}")
            return channel_id
        elif '/user/' in url or '/c/' in url:
            # Need to resolve the custom URL to get the channel ID
            username = url.rstrip('/').split('/')[-1]
            logging.debug(f"Resolving custom URL for username: {username}")
            # Use channels.list with forUsername
            request = youtube.channels().list(
                part='id',
                forUsername=username
            )
            response = request.execute()
            if 'items' in response and response['items']:
                channel_id = response['items'][0]['id']
                logging.debug(f"Resolved Channel ID using forUsername: {channel_id}")
                return channel_id
            else:
                # Try searching by custom URL
                logging.debug(f"forUsername did not return results for {username}. Trying search.")
                request = youtube.search().list(
                    part='snippet',
                    q=username,
                    type='channel',
                    maxResults=1
                )
                response = request.execute()
                if 'items' in response and response['items']:
                    channel_id = response['items'][0]['snippet']['channelId']
                    logging.debug(f"Resolved Channel ID using search: {channel_id}")
                    return channel_id
                else:
                    raise ValueError(f"Channel not found for URL: {url}")
        elif '/@' in url:
            # Handle @username URLs
            handle = url.rstrip('/').split('/')[-1]
            logging.debug(f"Resolving handle: {handle}")
            # Use search.list to find the channel by handle
            request = youtube.search().list(
                part='snippet',
                q=handle,
                type='channel',
                maxResults=1
            )
            response = request.execute()
            if 'items' in response and response['items']:
                channel_id = response['items'][0]['snippet']['channelId']
                logging.debug(f"Resolved Channel ID using handle: {channel_id}")
                return channel_id
            else:
                raise ValueError(f"Channel not found for URL: {url}")
        else:
            raise ValueError(f"Invalid YouTube channel URL format: {url}")
    except Exception as e:
        logging.error(f"Error extracting channel ID for URL {url}: {e}")
        raise e


def get_uploads_playlist_id(youtube, channel_id):
    """Retrieve the uploads playlist ID for a given channel."""
    try:
        request = youtube.channels().list(
            part='contentDetails,snippet',
            id=channel_id
        )
        response = request.execute()
        if 'items' in response and response['items']:
            uploads_playlist_id = response['items'][0]['contentDetails']['relatedPlaylists']['uploads']
            channel_title = response['items'][0]['snippet']['title']
            logging.debug(f"Uploads Playlist ID: {uploads_playlist_id} for Channel: {channel_title}")
            return uploads_playlist_id, channel_title
        else:
            raise ValueError(f"Channel not found for ID: {channel_id}")
    except Exception as e:
        logging.error(f"Error retrieving uploads playlist ID for channel ID {channel_id}: {e}")
        raise e


def get_all_video_ids(youtube, uploads_playlist_id):
    """Retrieve all video IDs from the uploads playlist."""
    video_ids = []
    next_page_token = None
    retries = 3

    while True:
        try:
            request = youtube.playlistItems().list(
                part='contentDetails',
                playlistId=uploads_playlist_id,
                maxResults=50,
                pageToken=next_page_token
            )
            response = request.execute()

            if 'items' not in response:
                break

            for item in response['items']:
                video_ids.append(item['contentDetails']['videoId'])

            next_page_token = response.get('nextPageToken')
            if not next_page_token:
                break
        except ssl.SSLError as ssl_err:
            logging.error(f"SSL Error: {ssl_err}. Retrying...")
            retries -= 1
            if retries > 0:
                logging.info(f"Retrying... Attempts left: {retries}")
                continue
            else:
                logging.critical("Max retries reached for SSL errors.")
                raise ssl_err
        except Exception as e:
            logging.error(f"An error occurred while fetching all video IDs: {e}")
            raise e

    logging.debug(f"Total Videos Retrieved (All): {len(video_ids)}")
    return video_ids


def get_last_n_video_ids(youtube, uploads_playlist_id, n=50):
    """Retrieve the last n video IDs from the uploads playlist."""
    video_ids = []
    next_page_token = None
    retries = 3

    while len(video_ids) < n:
        try:
            request = youtube.playlistItems().list(
                part='contentDetails',
                playlistId=uploads_playlist_id,
                maxResults=min(n - len(video_ids), 50),
                pageToken=next_page_token
            )
            response = request.execute()

            if 'items' not in response:
                break

            for item in response['items']:
                video_ids.append(item['contentDetails']['videoId'])
                if len(video_ids) == n:
                    break

            next_page_token = response.get('nextPageToken')
            if not next_page_token:
                break
        except ssl.SSLError as ssl_err:
            logging.error(f"SSL Error: {ssl_err}. Retrying...")
            retries -= 1
            if retries > 0:
                logging.info(f"Retrying... Attempts left: {retries}")
                continue
            else:
                logging.critical("Max retries reached for SSL errors.")
                raise ssl_err
        except Exception as e:
            logging.error(f"An error occurred while fetching last {n} video IDs: {e}")
            raise e

    logging.debug(f"Total Videos Retrieved (Last {n}): {len(video_ids)}")
    return video_ids


def get_videos_statistics(youtube, video_ids):
    """Retrieve statistics for a list of video IDs."""
    statistics = []
    # YouTube API allows up to 50 IDs per request
    for i in range(0, len(video_ids), 50):
        batch_ids = video_ids[i:i+50]
        try:
            request = youtube.videos().list(
                part='statistics, snippet',
                id=','.join(batch_ids)
            )
            response = request.execute()

            if 'items' in response:
                for item in response['items']:
                    video_id = item['id']
                    view_count = int(item['statistics'].get('viewCount', 0))
                    upload_date = item['snippet']['publishedAt']
                    title = item['snippet']['title']
                    statistics.append({
                        'video_id': video_id,
                        'title': title,
                        'view_count': view_count,
                        'upload_date': upload_date
                    })
        except ssl.SSLError as ssl_err:
            logging.error(f"SSL Error while fetching video statistics: {ssl_err}. Retrying...")
            retries = 3
            while retries > 0:
                try:
                    request = youtube.videos().list(
                        part='statistics, snippet',
                        id=','.join(batch_ids)
                    )
                    response = request.execute()

                    if 'items' in response:
                        for item in response['items']:
                            video_id = item['id']
                            view_count = int(item['statistics'].get('viewCount', 0))
                            upload_date = item['snippet']['publishedAt']
                            title = item['snippet']['title']
                            statistics.append({
                                'video_id': video_id,
                                'title': title,
                                'view_count': view_count,
                                'upload_date': upload_date
                            })
                    break  # Exit retry loop if successful
                except ssl.SSLError as ssl_err_retry:
                    retries -= 1
                    logging.error(f"Retry SSL Error: {ssl_err_retry}. Attempts left: {retries}")
                    if retries == 0:
                        logging.critical("Max retries reached for SSL errors during video statistics fetching.")
                        raise ssl_err_retry
                except Exception as e:
                    logging.error(f"An error occurred while fetching video statistics: {e}")
                    break
        except Exception as e:
            logging.error(f"An error occurred while fetching video statistics: {e}")
            continue

    logging.debug(f"Total Videos with Statistics Retrieved: {len(statistics)}")
    return statistics


class WorkerThread(QThread):
    progress = pyqtSignal(str)
    finished = pyqtSignal(object)

    def __init__(self, youtube, channel_url, n_videos):
        super().__init__()
        self.youtube = youtube
        self.channel_url = channel_url
        self.n_videos = n_videos

    def run(self):
        try:
            self.progress.emit(f"Processing channel: {self.channel_url}")
            channel_id = extract_channel_id(self.youtube, self.channel_url)
            self.progress.emit(f"Extracted Channel ID: {channel_id}")

            uploads_playlist_id, channel_title = get_uploads_playlist_id(self.youtube, channel_id)
            self.progress.emit(f"Channel Title: {channel_title}")

            if self.n_videos == "all":
                self.progress.emit("Fetching all video IDs...")
                video_ids = get_all_video_ids(self.youtube, uploads_playlist_id)
                self.progress.emit(f"Total Videos Retrieved: {len(video_ids)}")
            else:
                self.progress.emit("Fetching video IDs...")
                video_ids = get_last_n_video_ids(self.youtube, uploads_playlist_id, n=self.n_videos)
                self.progress.emit(f"Total Videos Retrieved: {len(video_ids)}")

            self.progress.emit("Fetching video statistics...")
            video_stats = get_videos_statistics(self.youtube, video_ids)
            self.progress.emit("Video statistics retrieved.")
            data = {
                'channel_title': channel_title,
                'video_stats': video_stats
            }
            self.finished.emit(data)
        except Exception as e:
            self.finished.emit({'error': str(e)})


class InteractiveChartView(QChartView):
    def __init__(self, chart, parent=None):
        super().__init__(chart, parent)

    def mouseDoubleClickEvent(self, event):
        self.chart().zoomReset()
        super().mouseDoubleClickEvent(event)


class YouTubeViewStatsApp(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("YouTube View Statistics Dashboard")
        self.setGeometry(100, 100, 1600, 900)
        self.layout = QVBoxLayout()

        # Dark theme setup
        self.setStyleSheet("background-color: #2E2E2E; color: #FFFFFF;")
        self.setPalette(self.create_dark_palette())

        # Splitter to divide input area and terminal
        splitter = QSplitter(Qt.Orientation.Horizontal)

        # Left Side: Controls
        left_widget = QWidget()
        left_layout = QVBoxLayout()

        # Channel URL Input
        self.channel_input_label = QLabel("Enter YouTube Channel URL:")
        left_layout.addWidget(self.channel_input_label)

        self.channel_input = QLineEdit()
        self.channel_input.setPlaceholderText("https://www.youtube.com/@YourChannelName")
        left_layout.addWidget(self.channel_input)

        # Dropdown and Analyze Button
        h_layout = QHBoxLayout()

        # Dropdown for selecting number of videos
        self.combo_label = QLabel("Select number of past videos:")
        h_layout.addWidget(self.combo_label)

        self.combo = QComboBox()
        # Populate dropdown with options: 50, 100, 150, ..., up to 1000 and "All"
        for i in range(50, 1050, 50):
            self.combo.addItem(str(i))
        self.combo.addItem("All")
        self.combo.setCurrentText("50")
        h_layout.addWidget(self.combo)

        # Analyze Button
        self.analyze_button = QPushButton("Analyze")
        self.analyze_button.setStyleSheet("""
            QPushButton {
                background-color: #2196F3;
                color: white;
                border-radius: 5px;
                padding: 5px 15px;
            }
            QPushButton:hover {
                background-color: #1976D2;
            }
        """)
        self.analyze_button.clicked.connect(self.analyze_channel)
        h_layout.addWidget(self.analyze_button)

        left_layout.addLayout(h_layout)

        # Export to Excel Button
        self.export_button = QPushButton("Export to Excel")
        self.export_button.setEnabled(False)  # Disabled until data is available
        self.export_button.setFixedSize(140, 40)
        self.export_button.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border-radius: 5px;
                padding: 5px 10px;
            }
            QPushButton:hover {
                background-color: #388E3C;
            }
        """)
        self.export_button.clicked.connect(self.export_to_excel)
        left_layout.addWidget(self.export_button)

        left_widget.setLayout(left_layout)

        # Right Side: Terminal-like Output
        right_widget = QWidget()
        right_layout = QVBoxLayout()

        # Terminal-like output
        self.terminal = QPlainTextEdit()
        self.terminal.setReadOnly(True)
        self.terminal.setFixedHeight(150)
        right_layout.addWidget(QLabel("Progress Output:"))
        right_layout.addWidget(self.terminal)

        right_widget.setLayout(right_layout)

        splitter.addWidget(left_widget)
        splitter.addWidget(right_widget)
        splitter.setStretchFactor(0, 3)
        splitter.setStretchFactor(1, 2)

        self.layout.addWidget(splitter)

        # Scroll area for charts
        self.scroll_area = QScrollArea()
        self.scroll_area_widget = QWidget()
        self.scroll_layout = QVBoxLayout()
        self.scroll_area_widget.setLayout(self.scroll_layout)
        self.scroll_area.setWidget(self.scroll_area_widget)
        self.scroll_area.setWidgetResizable(True)

        # Corrected alignment flag
        self.scroll_layout.setAlignment(Qt.AlignmentFlag.AlignTop)

        self.layout.addWidget(self.scroll_area)

        # Set the main layout
        self.setLayout(self.layout)

        # Initialize YouTube API client
        try:
            api_key = get_api_key()
            self.youtube = build('youtube', 'v3', developerKey=api_key)
            logging.info("YouTube API client initialized successfully.")
        except Exception as e:
            logging.critical(f"Failed to initialize YouTube API client: {e}")
            QMessageBox.critical(self, "Error", str(e))
            sys.exit()

        # List to keep track of worker threads
        self.worker_threads = []

        # Variables to store the latest DataFrame for export
        self.latest_df = None
        self.latest_channel_title = ""
        self.latest_n_videos = ""

    def create_dark_palette(self):
        palette = QPalette()
        palette.setColor(QPalette.ColorRole.Window, QColor("#2E2E2E"))
        palette.setColor(QPalette.ColorRole.WindowText, Qt.GlobalColor.white)
        palette.setColor(QPalette.ColorRole.Base, QColor("#1E1E1E"))
        palette.setColor(QPalette.ColorRole.AlternateBase, QColor("#3E3E3E"))
        palette.setColor(QPalette.ColorRole.ToolTipBase, Qt.GlobalColor.white)
        palette.setColor(QPalette.ColorRole.ToolTipText, Qt.GlobalColor.white)
        palette.setColor(QPalette.ColorRole.Text, Qt.GlobalColor.white)
        palette.setColor(QPalette.ColorRole.Button, QColor("#3E3E3E"))
        palette.setColor(QPalette.ColorRole.ButtonText, Qt.GlobalColor.white)
        palette.setColor(QPalette.ColorRole.BrightText, Qt.GlobalColor.red)
        palette.setColor(QPalette.ColorRole.Link, QColor("#2A82DA"))
        palette.setColor(QPalette.ColorRole.Highlight, QColor("#44475a"))
        palette.setColor(QPalette.ColorRole.HighlightedText, Qt.GlobalColor.white)
        return palette

    def log(self, message):
        self.terminal.appendPlainText(message)
        logging.info(message)

    def analyze_channel(self):
        channel_url = self.channel_input.text().strip()
        if not channel_url:
            QMessageBox.warning(self, "Input Required", "Please enter a YouTube channel URL.")
            return

        # Validate the URL format
        if not re.match(r'^https?://(www\.)?youtube\.com/', channel_url):
            QMessageBox.warning(self, "Invalid URL", "Please enter a valid YouTube channel URL.")
            return

        selected_value = self.combo.currentText()
        if selected_value.lower() == "all":
            n_videos = "all"
        else:
            try:
                n_videos = int(selected_value)
                if n_videos <= 0:
                    raise ValueError("Number of videos must be positive.")
            except ValueError as ve:
                logging.error(f"Invalid number of videos selected: {ve}")
                QMessageBox.critical(self, "Invalid Input", str(ve))
                return

        # Clear previous charts and logs
        self.clear_layout(self.scroll_layout)
        self.terminal.clear()

        # Disable the Export button until new data is available
        self.export_button.setEnabled(False)

        # Store the selected number of videos
        self.latest_n_videos = selected_value

        # Process the channel
        self.log(f"Starting analysis for {channel_url}")
        worker = WorkerThread(self.youtube, channel_url, n_videos)
        worker.progress.connect(self.log)
        worker.finished.connect(self.handle_worker_finished)
        self.worker_threads.append(worker)
        worker.start()

    def handle_worker_finished(self, data):
        worker = self.sender()
        if worker in self.worker_threads:
            self.worker_threads.remove(worker)

        if 'error' in data:
            self.log(f"Error: {data['error']}")
            QMessageBox.critical(self, "Error", data['error'])
            return

        channel_title = data['channel_title']
        video_stats = data['video_stats']

        self.latest_channel_title = channel_title

        if not video_stats:
            self.log(f"No video statistics available for {channel_title}.")
            QMessageBox.warning(self, "No Data", f"No video statistics available for {channel_title}.")
            return

        # Create a DataFrame
        df = pd.DataFrame(video_stats)
        # Convert upload_date to datetime
        df['upload_date'] = pd.to_datetime(df['upload_date'])
        # Remove timezone information if present
        if df['upload_date'].dt.tz is not None:
            df['upload_date'] = df['upload_date'].dt.tz_convert(None)
        # Sort by upload_date
        df.sort_values('upload_date', inplace=True)

        # Calculate moving average (window size of 5 or 10% of data length)
        window_size = max(1, len(df) // 10)  # Adjust window size based on data length
        df['moving_avg'] = df['view_count'].rolling(window=window_size).mean()

        self.latest_df = df.copy()  # Store the latest DataFrame for export

        # Plotting using PyQt6.QtCharts
        chart = QChart()
        chart.setTitle(f"View Counts for {channel_title}")
        chart.setAnimationOptions(QChart.AnimationOption.SeriesAnimations)

        # Series for View Count
        view_series = QLineSeries()
        view_series.setName("View Count")
        view_series.setPointsVisible(True)

        # Series for Moving Average
        avg_series = QLineSeries()
        avg_series.setName("Moving Average")
        avg_series.setPointsVisible(True)

        for index, row in df.iterrows():
            timestamp = row['upload_date'].timestamp() * 1000  # Convert to milliseconds
            view_series.append(timestamp, row['view_count'])
            if not pd.isna(row['moving_avg']):
                avg_series.append(timestamp, row['moving_avg'])

        chart.addSeries(view_series)
        chart.addSeries(avg_series)

        # Create axes
        axis_x = QDateTimeAxis()
        axis_x.setFormat("yyyy-MM-dd")
        axis_x.setTitleText("Upload Date")
        chart.addAxis(axis_x, Qt.AlignmentFlag.AlignBottom)
        view_series.attachAxis(axis_x)
        avg_series.attachAxis(axis_x)

        axis_y = QValueAxis()
        axis_y.setTitleText("View Count")
        chart.addAxis(axis_y, Qt.AlignmentFlag.AlignLeft)
        view_series.attachAxis(axis_y)
        avg_series.attachAxis(axis_y)

        # Enable tooltips using event filters
        view_series.hovered.connect(lambda point, state: self.show_tooltip(point, state, "View Count"))
        avg_series.hovered.connect(lambda point, state: self.show_tooltip(point, state, "Moving Average"))

        # Add legend
        chart.legend().setVisible(True)
        chart.legend().setAlignment(Qt.AlignmentFlag.AlignBottom)

        # Create a chart view and add it to the layout
        chart_view = InteractiveChartView(chart)
        chart_view.setRenderHint(QPainter.RenderHint.Antialiasing)

        # Enable zooming and panning
        chart_view.setRubberBand(QChartView.RubberBand.RectangleRubberBand)
        chart_view.setInteractive(True)

        # Add shadow effect for aesthetics
        shadow = QGraphicsDropShadowEffect()
        shadow.setBlurRadius(10)
        shadow.setXOffset(5)
        shadow.setYOffset(5)
        shadow.setColor(QColor("#000000"))
        chart_view.setGraphicsEffect(shadow)

        self.scroll_layout.addWidget(chart_view)

        # Enable the Export button now that data is available
        self.export_button.setEnabled(True)

    def show_tooltip(self, point, state, series_name):
        if state:
            try:
                date = datetime.fromtimestamp(point.x() / 1000).strftime("%Y-%m-%d")
                value = int(point.y())
                QToolTip.setFont(QFont('SansSerif', 10))
                QToolTip.showText(QCursor.pos(), f"{series_name} on {date}: {value} views", self)
            except Exception as e:
                logging.error(f"Error showing tooltip: {e}")

    def export_to_excel(self):
        if self.latest_df is None or self.latest_channel_title == "":
            QMessageBox.warning(self, "No Data", "There is no data to export. Please analyze a channel first.")
            return

        try:
            # Create a new Excel workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Video Statistics"

            # Write DataFrame to Excel
            for r in dataframe_to_rows(self.latest_df, index=False, header=True):
                ws.append(r)

            # Apply header styles
            header_font = Font(bold=True)
            for cell in ws[1]:
                cell.font = header_font

            # Create a Line Chart
            chart = LineChart()
            chart.title = f"View Counts for {self.latest_channel_title}"
            chart.style = 10
            chart.y_axis.title = 'View Count'
            chart.x_axis.title = 'Upload Date'

            # Define data for the chart
            data = Reference(ws, min_col=3, min_row=1, max_col=4, max_row=ws.max_row)
            cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.width = 20
            chart.height = 10

            # Add the chart to the sheet
            ws.add_chart(chart, "F2")

            # Generate filename
            current_date = datetime.now().strftime("%Y%m%d")
            sanitized_channel_title = re.sub(r'[\\/*?:"<>|]', "", self.latest_channel_title)
            filename = f"{sanitized_channel_title}-{self.latest_n_videos}-{current_date}.xlsx"

            # Save the workbook
            wb.save(filename)
            self.log(f"Data exported successfully to {filename}")
            QMessageBox.information(self, "Export Successful", f"Data exported successfully to {filename}")
        except Exception as e:
            logging.error(f"Error exporting to Excel: {e}")
            QMessageBox.critical(self, "Export Failed", f"An error occurred while exporting to Excel:\n{e}")

    def clear_layout(self, layout):
        if layout is not None:
            while layout.count():
                child = layout.takeAt(0)
                if child.widget() is not None:
                    child.widget().setParent(None)
                elif child.layout() is not None:
                    self.clear_layout(child.layout())

    def closeEvent(self, event):
        # Stop all running threads when the application is closed
        for worker in self.worker_threads:
            worker.quit()
            worker.wait()
        event.accept()


def main():
    app = QApplication(sys.argv)
    app.setStyle('Fusion')
    window = YouTubeViewStatsApp()
    window.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
